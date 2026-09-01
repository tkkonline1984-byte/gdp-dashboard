import io
import unittest
import zipfile

import fitz
import pandas as pd
from openpyxl import load_workbook

from converters import ConversionError, excel_to_pdf, images_to_excel, pdf_to_excel, pdf_to_jpg_zip


def sample_pdf() -> bytes:
    document = fitz.open()
    page = document.new_page()
    page.insert_text((72, 72), "Barcode 8850127000016 Product Test", fontsize=12)
    data = document.tobytes()
    document.close()
    return data


def sample_excel() -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(
            {"Barcode": ["8850127000016", "8850127000023"], "Product": ["A", "B"]}
        ).to_excel(writer, index=False, sheet_name="Products")
        pd.DataFrame({"Branch": ["Bangkok"]}).to_excel(writer, index=False, sheet_name="Branches")
    return output.getvalue()


class ConverterTests(unittest.TestCase):
    def test_pdf_to_jpg_zip(self):
        result = pdf_to_jpg_zip(sample_pdf(), dpi=120)
        self.assertEqual(result.items, 1)
        with zipfile.ZipFile(io.BytesIO(result.data)) as archive:
            self.assertEqual(archive.namelist(), ["page_001.jpg"])
            self.assertTrue(archive.read("page_001.jpg").startswith(b"\xff\xd8"))

    def test_pdf_text_to_excel(self):
        result = pdf_to_excel(sample_pdf())
        workbook = load_workbook(io.BytesIO(result.data), read_only=True)
        self.assertGreaterEqual(len(workbook.sheetnames), 1)
        values = " ".join(str(cell.value or "") for row in workbook.active.iter_rows() for cell in row)
        self.assertIn("8850127000016", values)

    def test_excel_to_pdf_keeps_all_sheets(self):
        result = excel_to_pdf(sample_excel(), "products.xlsx")
        self.assertTrue(result.data.startswith(b"%PDF"))
        self.assertEqual(result.items, 3)
        generated = fitz.open(stream=result.data, filetype="pdf")
        try:
            self.assertGreaterEqual(generated.page_count, 2)
        finally:
            generated.close()

    def test_rejects_broken_files(self):
        with self.assertRaises(ConversionError):
            pdf_to_jpg_zip(b"not-pdf")
        with self.assertRaises(ConversionError):
            excel_to_pdf(b"not-excel")

    def test_jpg_to_excel_when_ocr_is_available(self):
        import shutil
        if shutil.which("tesseract") is None:
            self.skipTest("Tesseract is not installed in this diagnostic runtime")
        from PIL import Image, ImageDraw, ImageFont

        image = Image.new("RGB", (900, 220), "white")
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 46)
        except OSError:
            font = ImageFont.load_default()
        ImageDraw.Draw(image).text((40, 70), "PRODUCT 8850127000016", fill="black", font=font)
        data = io.BytesIO()
        image.save(data, "PNG")
        result = images_to_excel([("product.png", data.getvalue())], language="eng")
        workbook = load_workbook(io.BytesIO(result.data), read_only=True)
        self.assertGreaterEqual(len(workbook.sheetnames), 1)
        values = " ".join(str(cell.value or "") for row in workbook.active.iter_rows() for cell in row)
        self.assertIn("8850127000016", values)


if __name__ == "__main__":
    unittest.main()
